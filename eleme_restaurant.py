#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
    @Time    : 2018-9-29 23:43
    @Author  : Joseph
    @File    : eleme_restaurant.py
    @Software: PyCharm
    @Description :
'''

import urllib.request
import os
import json
import time
from urllib import request

from openpyxl import Workbook, load_workbook

import setting
from geocoding import get_geocoding




def excelName(name):  # 根据日期生成文件
    targetDir = setting.RESTAURANT_INFO_DIR
    if not os.path.isdir(targetDir):
        os.mkdir(targetDir)
    excelName = name + "_饿了么_竞店_" + str(time.strftime("%Y-%m-%d") + ".xlsx")
    completePath = targetDir + "\\" + excelName
    return completePath


def write2Excel(jsondata, title):  # 根据不同的商圈地点写入数据，每个商圈地点占用excel 的一个sheet
    fileName = excelName(title)
    if (os.path.exists(fileName)):
        wb = load_workbook(fileName)
    else:
        wb = Workbook()
    if (wb.__contains__(title)):
        ws = wb[title]
        # ws.append([])
    else:
        ws = wb.create_sheet(title)
        ws.column_dimensions["A"].width = 10.0
        ws.column_dimensions["B"].width = 30.0
        ws.column_dimensions["C"].width = 10.0
        ws.column_dimensions["D"].width = 10.0
        ws.column_dimensions["E"].width = 10.0
        ws.column_dimensions["F"].width = 10.0
        ws.column_dimensions["G"].width = 20.0
        ws.column_dimensions["H"].width = 10.0
        ws.column_dimensions["I"].width = 10.0
        ws.column_dimensions["J"].width = 30.0
        ws.column_dimensions["K"].width = 10.0
        ws.column_dimensions["L"].width = 40.0
        ws.column_dimensions["M"].width = 10.0

        ws.append(["店铺ID", "店铺名称", "月销量", "起送价", "配送费", "平均送达速度", "营业时长", "评分", "评价数", "活动", "是否新开店", "店铺地址", "距离"])

    for i in range(len(jsondata)):
        attribute = ""
        row = jsondata[i]['restaurant']
        # print(type(row))
        # print(row)
        for activitie in row["activities"]:
            if "type" in activitie and activitie['type'] is 102:
                attribute = attribute + activitie["tips"]
            pass
        ws.append(
            [row["id"], row["name"], row["recent_order_num"], row["float_minimum_order_amount"], row["float_delivery_fee"], row["order_lead_time"], str(row["opening_hours"]), row["rating"], row["rating_count"], str(attribute), row["is_new"], row["address"], row["distance"]])

    wb.save(fileName)


def get_restaurant_data(latitude, longitude, limit, offset):
    req = urllib.request.Request(url=setting.ELEME_RESTAURANT_URL, headers=setting.DEFAULT_REQUEST_HEADERS)
    req.add_header("refer", "https://h5.ele.me/")  # 修改请求头的refer
    new_url = req.get_full_url()
    params = {
        "extras[]": "activities",
        "latitude": latitude,
        "longitude": longitude,
        "terminal": "h5",
        "limit": limit,
        "offset": offset
    }
    params = urllib.parse.urlencode(params)  # 请求参数编码
    req_full_url = new_url + params  # 重新生成url请求
    # print(req_full_url)
    web_page = request.urlopen(req_full_url)
    restaurant_data = json.loads(web_page.read().decode("utf-8"))
    return restaurant_data['items']


def get_restaurant_data_count(city, restaurant_name, count):
    offset = 0
    geocode_data = get_geocoding(city, restaurant_name)
    latitude = geocode_data['lat']
    longitude = geocode_data['lng']
    while offset + 30 <= count:
        data = get_restaurant_data(latitude, longitude, 30, offset)
        write2Excel(data, name)
        offset += 30
    if offset > count - 30:
        data = get_restaurant_data(latitude, longitude, count - offset, offset)
        write2Excel(data, name)


if __name__ == '__main__':  # 程序运行入口
    city = '北京'
    name = '益园文创基地'
    count = 100
    get_restaurant_data_count(city, name, count)







