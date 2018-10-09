#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
    @Time    : 2018-9-30 11:08
    @Author  : Joseph
    @File    : eleme_menu.py
    @Software: PyCharm
    @Description :
'''
import os
import json
import time
import requests

from openpyxl import Workbook, load_workbook

import eleme_restaurant_info
import setting



def excelName(name):  # 根据日期生成文件
    targetDir = setting.RESTAURANT_INFO_DIR
    if not os.path.isdir(targetDir):
        os.mkdir(targetDir)
    excelName = name + "_饿了么_菜单_" + str(time.strftime("%Y-%m-%d") + ".xlsx")
    completePath = targetDir + "\\" + excelName
    return completePath


def write2Excel(jsondata, title):
    fileName = excelName(title)
    if (os.path.exists(fileName)):
        wb = load_workbook(fileName)
    else:
        wb = Workbook()
    if (wb.__contains__(title)):
        ws = wb[title]
        # ws.append([])
    else:
        ws = wb.create_sheet(title)
        ws.column_dimensions["A"].width = 30.0
        ws.column_dimensions["B"].width = 10.0
        ws.column_dimensions["C"].width = 10.0
        ws.column_dimensions["D"].width = 10.0
        ws.column_dimensions["E"].width = 30.0
        ws.column_dimensions["F"].width = 10.0
        ws.column_dimensions["G"].width = 10.0
        ws.column_dimensions["H"].width = 10.0
        ws.append(["店铺名称", "菜品分类数", "SKU", "类别", "菜品名称", "月销售", "好评率", "价格"])

    for menus in (jsondata["menu"]):
        name = menus['name']
        if name == '热销':
            food = menus['foods']
            for i in range(len(food)):
                foods = food[i]
                if i == 0:
                    count = 0
                    for i in range(2, len(jsondata["menu"])):
                        count = count + len(jsondata["menu"][i]["foods"])
                    ws.append([jsondata["rst"]["name"], len(jsondata["menu"]) - 2, count,  "热销", foods["name"], foods["month_sales"], str(foods["satisfy_rate"]) + "%", foods["specfoods"][0]["price"]])
                else:
                    ws.append([jsondata["rst"]["name"], "", "", "热销", foods["name"], foods["month_sales"], str(foods["satisfy_rate"])+"%", foods["specfoods"][0]["price"]])
    wb.save(fileName)


# https://h5.ele.me/pizza/shopping/restaurants/161872698/batch_shop?extras=["activities","albums","license","identification","qualification"]&terminal=h5&latitude=39.952594&longitude=116.235477
def get_menu_data(restaurant_id, latitude, longitude):
    eleme_munu_url = setting.ELEME_MENU_URL % restaurant_id  # 占位符
    # req.add_header("refer", "https://h5.ele.me/shop/")
    # req.add_header("Cookie", "SID=D0hkTsFU4Lt9DQUhn0O85Qyxfa2rmIUw8dqg")
    # print(eleme_munu_url)
    params = {
        # "code":  random.random(),
        "extras": "[\"activities\",\"albums\",\"license\",\"identification\",\"qualification\"]",
        "latitude": latitude,
        "longitude": longitude,
        "terminal": "h5"
    }
    req = requests.request("GET", url=eleme_munu_url, headers=setting.DEFAULT_REQUEST_HEADERS, params=params)
    # params = urllib.parse.urlencode(params)  # 请求参数编码
    # req_full_url = eleme_munu_url + params  # 重新生成url请求
    # print(req_full_url)
    # web_page = request.urlopen(req_full_url)
    # print(web_page)
    # restaurant_data = json.loads(web_page.read().decode("utf-8"))
    restaurant_data = json.loads(req.text)
    print(restaurant_data)
    return restaurant_data


def get_all_rst_menu(fileName):
    if os.path.exists(fileName):
        wb = load_workbook(fileName)
    else:
        return
    for title in wb.sheetnames:
        ws = wb[title]
        for i in range(2, ws.max_row):
            params = {}  # 为生成请求参数做准备,店铺ID，经度，纬度
            params["restaurant_id"] = ws.cell(row=i, column=1).value
            params["latitude"] = ws.cell(row=i, column=2).value
            params["longitude"] = ws.cell(row=i, column=3).value
            menu_data = get_menu_data(params["restaurant_id"], params["latitude"], params["longitude"])
            write2Excel(menu_data, title)


if __name__ == '__main__':  # 程序运行入口
    offset = 0
    # data = get_menu_data(161872698, 39.952594, 116.235477)
    # write2Excel(data, data["rst"]["name"])
    excel_name = eleme_restaurant_info.excelName("益园文创基地")
    get_all_rst_menu(excel_name)
