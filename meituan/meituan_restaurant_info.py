#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
    @Time    : 2018-10-9 20:21
    @Author  : Joseph
    @File    : meituan_restaurant_info.py
    @Software: PyCharm
    @Description :
'''
import json
import os

import time

import requests
from openpyxl import load_workbook, Workbook

import setting
from utils.rearch_poi import get_poi


def excelName(name):  # 根据日期生成文件
    targetDir = setting.RESTAURANT_INFO_DIR
    if not os.path.isdir(targetDir):
        os.mkdir(targetDir)
    excelName = name + "_美团_竞店_" + str(time.strftime("%Y-%m-%d") + ".xlsx")
    completePath = targetDir + "\\" + excelName
    return completePath


def write2Excel(jsondata, title):
    fileName = excelName(title)
    if (os.path.exists(fileName)):
        wb = load_workbook(fileName)
    else:
        wb = Workbook()
    if (wb.__contains__(title)):
        ws = wb[title]
        # ws.append([])
    else:
        ws = wb.create_sheet(title)
        ws.column_dimensions["A"].width = 15.0
        ws.column_dimensions["B"].width = 35.0
        ws.column_dimensions["C"].width = 15.0
        ws.column_dimensions["D"].width = 15.0
        ws.column_dimensions["E"].width = 15.0
        ws.column_dimensions["F"].width = 15.0
        ws.column_dimensions["G"].width = 10.0
        ws.column_dimensions["G"].width = 10.0

        ws.append(["店铺ID", "店铺名称", "起送价", "配送费", "人均消费", "平均送达速度", "评分", "距离"])

    for i in range(len(jsondata)):
        row = jsondata[i]
        print(row)
        ws.append(
            [row["mtWmPoiId"], row["shopName"], row["minPriceTip"], row["shippingFeeTip"], row["averagePriceTip"], row["deliveryTimeTip"], row["wmPoiScore"]/10, row["distance"]])

    wb.save(fileName)


def get_restaurant_data(index, sort_id, latitude, longitude):
    meituan_restaurant_url = "http://m.waimai.meituan.com/waimai/ajax/newm/kingkongshoplist"
    headers = {
        "Cookie": "iuuid=1049974659455627100; waimai_cityname=%E5%9F%8E%E5%B8%82%E5%90%8D%E5%B7%B2%E5%88%A0%E9%99%A4",
        "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1"
    }
    params = {
        "startIndex": index,
        "sortId": sort_id,
        "navigateType": 101260,
        "firstCategoryId": 101260,
        "secondCategoryId": 101262,
        "initialLat": latitude,
        "initialLng": longitude,
        "geoType": 2
    }
    req = requests.request("GET", url=meituan_restaurant_url, headers=headers, params=params)
    restaurant_data = json.loads(req.text)
    # print(restaurant_data)
    return restaurant_data["data"]["shopList"]


def get_restaurant_data_count(keyword, count, sort_id):
    poi_data = get_poi(keyword)
    latitude = poi_data['latitude']
    longitude = poi_data['longitude']
    index = 0
    while (index + 1)*20 <= count:
        data = get_restaurant_data(index, sort_id, latitude, longitude)
        write2Excel(data, keyword)
        index += 1
    # if (index + 1)*20 > count - 20:
    #     data = get_restaurant_data(index, sort_id, latitude, longitude)
    #     write2Excel(data, keyword)


if __name__ == '__main__':  # 程序运行入口
    keyword = "北京益园文创基地"
    count = 20  # 20的整倍数
    sortId = 1  # 销量最高
    # sortId = 5  # 距离最近
    secondCategoryId = 101262  # 快餐便当
    get_restaurant_data_count(keyword, count, sortId)

