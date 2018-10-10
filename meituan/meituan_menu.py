#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
    @Time    : 2018-10-10 19:02
    @Author  : Joseph
    @File    : meituan_menu.py
    @Software: PyCharm
    @Description :
'''

import os
import json
import time
import requests

from openpyxl import Workbook, load_workbook

from eleme import eleme_restaurant_info
import setting
from meituan import meituan_restaurant_info


def excelName(name):  # 根据日期生成文件
    targetDir = setting.RESTAURANT_INFO_DIR
    if not os.path.isdir(targetDir):
        os.mkdir(targetDir)
    excelName = name + "_美团_菜单_" + str(time.strftime("%Y-%m-%d") + ".xlsx")
    completePath = targetDir + "\\" + excelName
    return completePath


def write2Excel(jsondata, title):
    fileName = excelName(title)
    if (os.path.exists(fileName)):
        wb = load_workbook(fileName)
    else:
        wb = Workbook()
    if (wb.__contains__(title)):
        ws = wb[title]
        # ws.append([])
    else:
        ws = wb.create_sheet(title)
        ws.column_dimensions["A"].width = 25.0
        ws.column_dimensions["B"].width = 10.0
        ws.column_dimensions["C"].width = 10.0
        ws.column_dimensions["D"].width = 15.0
        ws.column_dimensions["E"].width = 15.0
        ws.column_dimensions["F"].width = 10.0
        ws.column_dimensions["G"].width = 15.0
        ws.column_dimensions["H"].width = 10.0
        ws.column_dimensions["I"].width = 10.0
        ws.column_dimensions["J"].width = 10.0
        ws.column_dimensions["K"].width = 10.0
        ws.append(["店铺名称", "菜品分类数", "SKU数量", "配送服务", "营业时间", "类别", "菜品名称", "月销售", "点赞", "产品标格", "实付价格"])

    for menus in (jsondata["categoryList"]):
        name = menus['categoryName']
        if name == '热销':
            food = menus['spuList']
            for i in range(len(food)):
                foods = food[i]
                if i == 0:
                    count = 0
                    for i in range(1, len(jsondata["categoryList"])):
                        count = count + len(jsondata["categoryList"][i]["spuList"])
                    ws.append([jsondata["shopInfo"]["shopName"], len(jsondata["categoryList"]) - 1, count, jsondata["shopInfo"]["deliveryMsg"], jsondata["shopInfo"]["shippingTime"], name, foods["spuName"], foods["sellStatus"], foods["praiseNum"], foods["originPrice"], foods["currentPrice"]])
                else:
                    ws.append(["", "", "", "", "", name, foods["spuName"], foods["sellStatus"], foods["praiseNum"], foods["originPrice"], foods["currentPrice"]])
    wb.save(fileName)


def get_menu_data(restaurant_id, latitude, longitude):
    eleme_munu_url = "https://m.waimai.meituan.com/waimai/ajax/newm/menu"
    headers = {
        "Cookie": "iuuid=1049974659455627326; waimai_cityname=%E5%9F%8E%E5%B8%82%E5%90%8D%E5%B7%B2%E5%88%A0%E9%99%A4",
        "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1"
    }
    params = {
        "mtWmPoiId": restaurant_id,
        # "initialLat": latitude,
        # "initialLng": longitude,
        "geoType": 2,
        "dpShopId": -1
    }
    req = requests.request("GET", url=eleme_munu_url, headers=headers, params=params)
    menu_data = json.loads(req.text)
    print(menu_data)
    return menu_data["data"]


def get_all_rst_menu(fileName):
    if os.path.exists(fileName):
        wb = load_workbook(fileName)
    else:
        return
    for title in wb.sheetnames:
        ws = wb[title]
        for i in range(2, ws.max_row):
            params = {}  # 为生成请求参数做准备,店铺ID，经度，纬度
            params["restaurant_id"] = ws.cell(row=i, column=1).value
            menu_data = get_menu_data(params["restaurant_id"], "", "")
            write2Excel(menu_data, title)


if __name__ == '__main__':  # 程序运行入口
    offset = 0
    # data = get_menu_data(161872698, 39.952594, 116.235477)
    # write2Excel(data, data["rst"]["name"])
    excel_name = meituan_restaurant_info.excelName("北京益园文创基地")
    get_all_rst_menu(excel_name)
